import paramiko
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart
import numpy as np
from openpyxl.chart import (
    PieChart,
    Reference
)
from openpyxl.chart.series import DataPoint


class Os_info:

    def __init__(self, username, password, hosts):
        self.username = username
        self.password = password
        self.hosts = hosts

    def check_disk(self):
        wb = Workbook()
        ws = wb.active
        #ws = wb.create_sheet("disk_info")
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        start_time = time.time()
        ws.append(["节点", "使用率(%)", "目录", "节点磁盘阈值高于80%数量"])
        i = 2
        n = 0
        start_num = 0
        end_num = 0
        hosts_list = []
        for host in self.hosts:
            hosts_list.append(host)
            client.connect(hostname=host, port=22, username=self.username, password=self.password)
            stdin, stdout, stderr = client.exec_command("ip=`ip a  | grep inet | awk '{if(NR==3) print $0}' | sed -e 's/\/.*//g' | awk '{print $2}'`;df -h | grep -v Filesystem | sed -e \'s/%//g\' | awk -v ip=$ip \'{if($5>=80) print ip,$5,$6}\'")
            disk_info = stdout.read().decode().split()
            disk_info1 = np.array(disk_info)
            disk_info1 = disk_info1.reshape(-1, 3)
            disk_num = len(disk_info1)
            #print(disk_num)
            for disk in range(0, len(disk_info1)):
                #disk_info1[disk].append(disk_num)
                a = list(disk_info1[disk])
                a.append(int(disk_num))
                #print(a[0] == host)
                #print(disk_info1[disk])
                ws.append(a)
            num_list = []
            # 从第几行开始
            while True:
                r = ws.cell(i, 1).value
                if r:
                    num_list.append(r)
                else:
                    ws.delete_rows(i)
                    break
                i += 1
            flag = num_list[0]
            print(flag)
            #print(num_list[0])
            print(num_list)
            start = 0
            end = 0
            if n >= len(num_list) - 1:
                for n in range(len(num_list)):
                    if n == len(num_list) - 1:
                        end_num = start_num + len(num_list) + len(num_list) - 1
                        start_num = start_num + len(num_list)
                        ws.merge_cells("A" + str(start_num + 2) + ":A" + str(end_num + 2))
                        ws.merge_cells("D" + str(start_num + 2) + ":D" + str(end_num + 2))
            else:
                for n in range(len(num_list)):
                    if num_list[n] != flag:
                        end = n - 1
                        if end >= start:
                            ws.merge_cells("A" + str(start + 2) + ":A" + str(end + 2))
                            ws.merge_cells("D" + str(start + 2) + ":D" + str(end + 2))
                            start = end + 1
                    if n == len(num_list) - 1:
                        end = n
                        ws.merge_cells("A" + str(start + 2) + ":A" + str(end + 2))
                        ws.merge_cells("D" + str(start + 2) + ":D" + str(end + 2))
            sh_name = wb.sheetnames
            sh = wb[sh_name[0]]
            sh.title = "disk_info"
            print(sh.title)
            #disk_sum = (disk - 1) * (len(disk_info1))
            #ws["D" + str(disk_sum)] = disk_sum
        for num, ip in enumerate(hosts_list):
            ws["J"+str(num+1)] = ip
        pie = PieChart()
        labels = Reference(ws, min_col=10, min_row=1, max_row=len(num_list)*10+1)
        data = Reference(ws, min_col=4, min_row=1, max_row=len(num_list)*10+1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "节点磁盘信息"
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]
        ws.add_chart(pie, "F1")
        ws.column_dimensions.group("x", hidden=True)
        wb.save('disk_info.xlsx')
        end_time = time.time()
        print(end_time - start_time)
        #print(wb.sheetnames)
        client.close()
        wb.close()

    def check_mem(self):
        #wb1 = Workbook()
        #ws1 = wb1.active
        ws1 = load_workbook('disk_info.xlsx')
        ws2 = ws1.create_sheet('memory_info')
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ws2.append(["节点", "内存已使用(%)"])
        mem_sum = 1
        for host in self.hosts:
            client.connect(hostname=host, port=22, username=self.username, password=self.password)
            stdin, stdout, stderr = client.exec_command("ip=`ip a  | grep inet | awk '{if(NR==3) print $0}' | sed -e 's/\/.*//g' | awk '{print $2}'`;free -m |awk '{if(NR==2) print $0}'|awk -v ip=$ip '{printf \"%s,%.2f\",ip,100-$7/$2*100}'")
            mem = stdout.read().decode().split(',')
            for num, info in enumerate(mem):
                if num == 1:
                    mem1 = float(mem[num])
                    mem.remove(mem[num])
                    mem.append(mem1)
            mem_sum += 1
            ws2.append(mem)
        print(mem_sum)
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10.0
        chart1.title = "节点内存已使用率"
        chart1.y_axis.title = '内存已使用数值'
        chart1.x_axis.title = '节点IP'
        data = Reference(ws2, min_col=1, min_row=1, max_row=mem_sum, max_col=2)
        cats = Reference(ws2, min_row=2, min_col=1, max_row=mem_sum)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws2.add_chart(chart1, "D5")
        ws1.save('disk_info.xlsx')
        client.close()
        ws1.close()

    def check_cpu(self):
        ws1 = load_workbook('disk_info.xlsx')
        ws2 = ws1.create_sheet('cpu_info')
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ws2.append(["节点", "cpu已消耗(%)"])
        for host in self.hosts:
            client.connect(hostname=host, port=22, username=self.username, password=self.password)
            stdin, stdout, stderr = client.exec_command("ip=`ip a  | grep inet | awk '{if(NR==3) print $0}' | sed -e 's/\/.*//g' | awk '{print $2}'`;vmstat |awk -v ip=$ip 'NR==3 {sum=$13+$14;print ip,sum}'")
            cpus = stdout.read().decode().split()
            print(cpus)
            for num, cpu in enumerate(cpus):
                if num == 1:
                    cpu1 = int(cpus[num])
                    cpus.remove(cpus[num])
                    cpus.append(cpu1)
            ws2.append(cpus)
            chart1 = BarChart()
            chart1.type = "col"
            chart1.style = 10.0
            chart1.title = "节点cpu已消耗"
            chart1.y_axis.title = 'cpu消耗数值'
            chart1.x_axis.title = '节点IP'
            data = Reference(ws2, min_col=1, min_row=1, max_row=10, max_col=2)
            cats = Reference(ws2, min_row=2, min_col=1, max_row=10)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.shape = 4
            ws2.add_chart(chart1, "D5")
            ws1.save('disk_info.xlsx')
            client.close()
            ws1.close()


if __name__ == '__main__':

    username = "root"
    password = "123456"
    hosts = []
    with open('ip', 'r') as f:
        line = f.readline().strip()
        while line:
            hosts.append(line)
            line = f.readline().strip()
    print(hosts)
    #hosts = ["10.211.55.7", "10.211.55.8", "10.211.55.9", "10.211.55.9", "10.211.55.8", "10.211.55.7", "10.211.55.8", "10.211.55.9", "10.211.55.9", "10.211.55.8"]
    o1 = Os_info(username, password, hosts)
    o1.check_disk()
    o1.check_mem()
    o1.check_cpu()